from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import csv
import logging


class BasicSheet(object):
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(self.file_name)
        self.active_sheet = self.wb.active

    def write_to_sheet(self, row, column, data):
        """assign value <data> to the cell in given row and column"""
        self.active_sheet.cell(row=row,
                               column=self._get_column_index_or_default(
                                column), value=data)
        self.wb.save(self.file_name)

    def write_row(self, row, data):
        """given a row index and a list of data points, fill the row
        with all elements in the list of data
        """
        i = 1
        for d in data:
            self.write_to_sheet(row, i, d)
            i += 1

    def write_next_row(self, data):
        """given a list of data points, fill the next unused row with
        all elements in the list of data
        """
        self.write_row(self.get_next_row_index(), data)

    def set_active_sheet(self, sheet_name):
        """set the active sheet to the given sheet_name, and
        create a new sheet if it doesn't already exist
        """
        if sheet_name not in self.wb:
            logging.info('Sheet %s does not exist. Creating one now and\
             setting as active sheet...', sheet_name)
            self.wb.create_sheet(title=sheet_name)
        self.active_sheet = self.wb[sheet_name]

    def get_header_row_data(self):
        """return a generator for the first row in the active sheet"""
        return self.get_row_data(min_row=1, max_row=1)

    def get_row_data(self, min_row=None, max_row=None):
        """given a min and max row, return a generator for all of the
        requested rows
        """
        return self.active_sheet.iter_rows(min_row=min_row, max_row=max_row)

    def get_all_sheets(self):
        """return the name of all the existing sheets"""
        return self.wb.sheetnames

    def get_next_row_index(self):
        """return the index of the next unused row"""
        return self.active_sheet.max_row + 1

    def get_next_col_index(self):
        """return the index of the next unused col"""
        return self.active_sheet.max_column + 1

    def _get_column_index_or_default(self, column):
        """given a column index, return an integer equivalent if its not
        already and integer
        """
        if isinstance(column, int):
            return column
        return column_index_from_string(column)

    def _iter_cells(self, row):
        """given a row generator element, return an array of all
        cell values within the row
        """
        return [str(cell.internal_value) for cell in row]
